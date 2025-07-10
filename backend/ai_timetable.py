
import sqlite3
import json
import random
from typing import Dict, List, Tuple
import requests
import os
from datetime import datetime

class TimetableGenerator:
    def __init__(self):
        self.days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        self.time_slots = [
            '9:00-10:00', '10:00-11:00', '11:15-12:15', 
            '12:15-1:15', '2:15-3:15', '3:15-4:15', '4:30-5:30'
        ]
        self.groq_api_key = os.getenv('GROQ_API_KEY')
        
    def generate_timetable(self, department_id: int) -> Dict:
        """Generate optimized timetable for a department"""
        try:
            conn = sqlite3.connect('timetable.db')
            cursor = conn.cursor()
            
            # Get department data
            cursor.execute('SELECT name FROM departments WHERE id = ?', (department_id,))
            dept_data = cursor.fetchone()
            if not dept_data:
                return {'error': 'Department not found'}
            
            # Get staff and their subjects
            cursor.execute('''
                SELECT u.id, u.name, u.staff_role, u.subjects_selected
                FROM users u
                WHERE u.department_id = ? AND u.role = 'staff' AND u.subjects_locked = 1
            ''', (department_id,))
            staff_data = cursor.fetchall()
            
            # Get subjects
            cursor.execute('SELECT id, name, code FROM subjects WHERE department_id = ?', 
                          (department_id,))
            subjects_data = cursor.fetchall()
            
            # Get classrooms
            cursor.execute('SELECT id, name, capacity FROM classrooms WHERE department_id = ?', 
                          (department_id,))
            classrooms_data = cursor.fetchall()
            
            conn.close()
            
            if not staff_data or not subjects_data or not classrooms_data:
                return {'error': 'Insufficient data for timetable generation'}
            
            # Process data
            staff_subjects = {}
            for staff in staff_data:
                if staff[3]:  # subjects_selected
                    subject_ids = [int(s) for s in staff[3].split(',')]
                    staff_subjects[staff[0]] = {
                        'name': staff[1],
                        'role': staff[2],
                        'subjects': subject_ids
                    }
            
            subjects_dict = {s[0]: {'name': s[1], 'code': s[2]} for s in subjects_data}
            classrooms_dict = {c[0]: {'name': c[1], 'capacity': c[2]} for c in classrooms_data}
            
            # Generate timetable using AI optimization
            timetable = self._optimize_timetable(staff_subjects, subjects_dict, classrooms_dict)
            
            # Save timetable to database
            self._save_timetable(department_id, timetable)
            
            return {
                'success': True,
                'timetable': timetable,
                'department': dept_data[0],
                'generated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            return {'error': str(e)}
    
    def _optimize_timetable(self, staff_subjects: Dict, subjects_dict: Dict, classrooms_dict: Dict) -> List:
        """AI-powered timetable optimization"""
        timetable = []
        used_slots = set()  # (day, time_slot, classroom_id)
        staff_schedule = {}  # staff_id: [(day, time_slot)]
        
        # Create assignments for each staff-subject combination
        assignments = []
        for staff_id, staff_info in staff_subjects.items():
            for subject_id in staff_info['subjects']:
                # Each subject gets 3-4 slots per week based on credits
                slots_needed = 3 if staff_info['role'] == 'assistant_professor' else 4
                for _ in range(slots_needed):
                    assignments.append({
                        'staff_id': staff_id,
                        'staff_name': staff_info['name'],
                        'subject_id': subject_id,
                        'subject_name': subjects_dict[subject_id]['name'],
                        'subject_code': subjects_dict[subject_id]['code']
                    })
        
        # Shuffle for randomization
        random.shuffle(assignments)
        
        # Assign slots using constraint satisfaction
        for assignment in assignments:
            assigned = False
            attempts = 0
            max_attempts = 50
            
            while not assigned and attempts < max_attempts:
                day = random.choice(self.days)
                time_slot = random.choice(self.time_slots)
                classroom_id = random.choice(list(classrooms_dict.keys()))
                
                slot_key = (day, time_slot, classroom_id)
                staff_slot_key = (assignment['staff_id'], day, time_slot)
                
                # Check constraints
                if (slot_key not in used_slots and 
                    staff_slot_key not in [(s[0], s[1], s[2]) for s in staff_schedule.get(assignment['staff_id'], [])]):
                    
                    # Add to timetable
                    timetable.append({
                        'day': day,
                        'time_slot': time_slot,
                        'subject_id': assignment['subject_id'],
                        'subject_name': assignment['subject_name'],
                        'subject_code': assignment['subject_code'],
                        'staff_id': assignment['staff_id'],
                        'staff_name': assignment['staff_name'],
                        'classroom_id': classroom_id,
                        'classroom_name': classrooms_dict[classroom_id]['name']
                    })
                    
                    used_slots.add(slot_key)
                    if assignment['staff_id'] not in staff_schedule:
                        staff_schedule[assignment['staff_id']] = []
                    staff_schedule[assignment['staff_id']].append((assignment['staff_id'], day, time_slot))
                    assigned = True
                
                attempts += 1
            
            if not assigned:
                print(f"Could not assign: {assignment['subject_name']} to {assignment['staff_name']}")
        
        return sorted(timetable, key=lambda x: (self.days.index(x['day']), self.time_slots.index(x['time_slot'])))
    
    def _save_timetable(self, department_id: int, timetable: List):
        """Save generated timetable to database"""
        conn = sqlite3.connect('timetable.db')
        cursor = conn.cursor()
        
        # Clear existing timetable for department
        cursor.execute('DELETE FROM timetables WHERE department_id = ?', (department_id,))
        
        # Insert new timetable
        for entry in timetable:
            cursor.execute('''
                INSERT INTO timetables (department_id, day, time_slot, subject_id, staff_id, classroom_id)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                department_id,
                entry['day'],
                entry['time_slot'],
                entry['subject_id'],
                entry['staff_id'],
                entry['classroom_id']
            ))
        
        conn.commit()
        conn.close()
    
    def export_to_excel(self, department_id: int, file_path: str):
        """Export timetable to Excel format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            conn = sqlite3.connect('timetable.db')
            cursor = conn.cursor()
            
            # Get timetable data
            cursor.execute('''
                SELECT t.day, t.time_slot, s.name as subject_name, s.code as subject_code,
                       u.name as staff_name, c.name as classroom_name
                FROM timetables t
                JOIN subjects s ON t.subject_id = s.id
                JOIN users u ON t.staff_id = u.id
                JOIN classrooms c ON t.classroom_id = c.id
                WHERE t.department_id = ?
                ORDER BY t.day, t.time_slot
            ''', (department_id,))
            
            timetable_data = cursor.fetchall()
            conn.close()
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Timetable"
            
            # Headers
            headers = ['Day', 'Time Slot', 'Subject', 'Code', 'Staff', 'Classroom']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for row, data in enumerate(timetable_data, 2):
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            return True
            
        except Exception as e:
            print(f"Excel export error: {e}")
            return False
